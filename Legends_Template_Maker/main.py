#!/usr/bin/env python3
"""
Legend Table Image Processor — QS AUTOMATION
=============================================
Automatically crops each symbol from an engineering-drawing legend table
and names the output file from the OCR-read description text.

Usage:
    python legend_processor.py <image_path>

Example:
    python legend_processor.py legend.jpeg
"""

import os
import re
import sys
import tempfile

import cv2
import numpy as np
import pytesseract
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.isfile(TESSERACT_CMD):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(_SCRIPT_DIR, "output_icons")
SYMBOL_PADDING = 5
HEADER_KEYWORDS = {"legend", "symbol", "description"}
MIN_ROW_HEIGHT = 10

# Upscale factor for the entire image before processing.
# Higher values give much better OCR and line detection accuracy.
TARGET_WIDTH = 2400

# OCR ROI upscale — each description cell is further upscaled by this factor
# before being fed to Tesseract for maximum text clarity.
OCR_UPSCALE_FACTOR = 3


# ---------------------------------------------------------------------------
# Common OCR correction dictionary for engineering legend text
# ---------------------------------------------------------------------------
OCR_CORRECTIONS = {
    # Single-word fixes
    "ughting": "lighting",
    "lghting": "lighting",
    "lghtins": "lighting",
    "lishting": "lighting",
    "lighitng": "lighting",
    "lighiing": "lighting",
    "lighling": "lighting",
    "lightng": "lighting",
    "lichtng": "lighting",
    "lichting": "lighting",
    "cchtrol": "control",
    "ochtrol": "control",
    "cohtrol": "control",
    "contrcl": "control",
    "contro": "control",
    "controi": "control",
    "contrpl": "control",
    "cortrol": "control",
    "conrol": "control",
    "contol": "control",
    "contrl": "control",
    "uchtins": "lighting",
    "eedside": "bedside",
    "sedroomentry": "bedroom entry",
    "eedroch": "bedroom",
    "bedroch": "bedroom",
    "bedrcom": "bedroom",
    "bedroon": "bedroom",
    "sedroom": "bedroom",
    "poder": "powder",
    "roca": "room",
    "roch": "room",
    "roou": "room",
    "acch": "room",
    "roou": "room",
    "stasf": "staff",
    "stalf": "staff",
    "mt_chen": "kitchen",
    "kitcben": "kitchen",
    "kitehen": "kitchen",
    "kiichen": "kitchen",
    "kichen": "kitchen",
    "ktchen": "kitchen",
    "lftloesy": "lift lobby",
    "lftlobby": "lift lobby",
    "liftloeby": "lift lobby",
    "liftlobey": "lift lobby",
    "liftiobby": "lift lobby",
    "closencloset": "closet/closet",
    "man": "main",
    "corrdor": "corridor",
    "coridor": "corridor",
    "corridcr": "corridor",
    "corridar": "corridor",
    "corridor": "corridor",
    "hryac": "hvac",
    "hvag": "hvac",
    "h_ac": "hvac",
    "hac": "hvac",
    "hy ac": "hvac",
    "hvec": "hvac",
    "curtan": "curtain",
    "curtian": "curtain",
    "curiain": "curtain",
    "curtam": "curtain",
    "curtaln": "curtain",
    "pr": "pir",
    "fir": "pir",
    "senscr": "sensor",
    "sensot": "sensor",
    "sensof": "sensor",
    "sensdr": "sensor",
    "senscor": "sensor",
    "sensar": "sensor",
    "docr": "door",
    "doar": "door",
    "doot": "door",
    "pahel": "panel",
    "ppel": "panel",
    "pnel": "panel",
    "p-nel": "panel",
    "pabel": "panel",
    "panei": "panel",
    "panet": "panel",
    "panef": "panel",
    "pane1": "panel",
    "atomtion": "automation",
    "automtion": "automation",
    "automaton": "automation",
    "automalion": "automation",
    "gooyeouksouis": "600x800x150mm",
    "rexcing": "reading",
    "reacing": "reading",
    "readng": "reading",
    "readinig": "reading",
    "lifng": "living",
    "lving": "living",
    "liying": "living",
    "lvng": "living",
    "scehe": "scene",
    "sene": "scene",
    "scere": "scene",
    "scane": "scene",
    "led": "media",
    "loeby": "lobby",
    "staircase": "staircase",
    "lght": "light",
    "lighit": "light",
    "lghi": "light",
    "ligiht": "light",
    "thermostat": "thermostat",
    "thermcstat": "thermostat",
    "thermosfat": "thermostat",
    "exhausi": "exhaust",
    "exhausl": "exhaust",
    "bathrom": "bathroom",
    "bathrcom": "bathroom",
    "balhroom": "bathroom",
    "baihroom": "bathroom",
    "entrancé": "entrance",
    "entranee": "entrance",
    "enirance": "entrance",
    "entiance": "entrance",
    "vic": "wic",
    "vac": "wic",
    "lounge": "lounge",
    "launfe": "lounge",
    "launce": "lounge",
    "games": "games",
    "gemes": "games",
    "ganes": "games",
    "houe": "home",
    "hone": "home",
    "nome": "home",
    "horre": "home",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sanitize_filename(text: str) -> str:
    """Turn OCR text into a safe, lowercase, underscore-separated filename."""
    text = text.strip().lower()
    text = re.sub(r"[\s/\\:*?\"<>|]+", "_", text)
    text = re.sub(r"[^a-z0-9_\-]", "", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def unique_filename(name: str, used: dict) -> str:
    """Return a unique name, appending _1, _2, … on collision."""
    if name not in used:
        used[name] = 0
        return name
    used[name] += 1
    new = f"{name}_{used[name]}"
    while new in used:
        used[name] += 1
        new = f"{name}_{used[name]}"
    used[new] = 0
    return new


def upscale_image(img: np.ndarray, target_width: int) -> tuple[np.ndarray, float]:
    """Upscale image so width = target_width. Returns (scaled_img, scale_factor)."""
    h, w = img.shape[:2]
    if w >= target_width:
        return img, 1.0
    scale = target_width / w
    new_w = int(w * scale)
    new_h = int(h * scale)
    scaled = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_CUBIC)
    return scaled, scale


def apply_ocr_corrections(text: str) -> str:
    """
    Apply dictionary-based corrections to fix common OCR mistakes
    in engineering legend descriptions.
    """
    words = text.lower().split()
    corrected = []
    for word in words:
        # Strip punctuation for matching but preserve structure
        clean = re.sub(r'[^a-z0-9/]', '', word)
        if clean in OCR_CORRECTIONS:
            corrected.append(OCR_CORRECTIONS[clean])
        else:
            corrected.append(word)
    return " ".join(corrected)


# ---------------------------------------------------------------------------
# Grid / line detection
# ---------------------------------------------------------------------------

def build_line_mask(img_bgr: np.ndarray) -> np.ndarray:
    """
    Create a binary mask combining red lines and dark/black lines.
    """
    hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)

    # Red hue (wraps around 0/180)
    red1 = cv2.inRange(hsv, np.array([0, 40, 40]), np.array([10, 255, 255]))
    red2 = cv2.inRange(hsv, np.array([165, 40, 40]), np.array([180, 255, 255]))
    red_mask = red1 | red2

    # Dark / black lines
    dark_mask = cv2.inRange(hsv, np.array([0, 0, 0]), np.array([180, 255, 80]))

    combined = cv2.bitwise_or(red_mask, dark_mask)
    return combined


def detect_lines(line_mask: np.ndarray):
    """
    Extract sorted lists of unique horizontal Y and vertical X coordinates
    using morphological operations.
    """
    h, w = line_mask.shape[:2]

    h_kernel_len = max(w // 20, 30)
    v_kernel_len = max(h // 20, 30)

    # Horizontal lines
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (h_kernel_len, 1))
    h_lines = cv2.morphologyEx(line_mask, cv2.MORPH_OPEN, h_kernel, iterations=2)
    h_proj = np.where(h_lines.max(axis=1) > 0)[0]
    h_coords = _cluster_coords(h_proj, gap=max(h // 150, 3))

    # Vertical lines
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_kernel_len))
    v_lines = cv2.morphologyEx(line_mask, cv2.MORPH_OPEN, v_kernel, iterations=2)
    v_proj = np.where(v_lines.max(axis=0) > 0)[0]
    v_coords = _cluster_coords(v_proj, gap=max(w // 150, 3))

    return h_coords, v_coords


def _cluster_coords(coords: np.ndarray, gap: int = 5) -> list[int]:
    """Group nearby coordinates, return median of each cluster."""
    if len(coords) == 0:
        return []
    clusters: list[list[int]] = [[coords[0]]]
    for c in coords[1:]:
        if c - clusters[-1][-1] <= gap:
            clusters[-1].append(c)
        else:
            clusters.append([c])
    return [int(np.median(cl)) for cl in clusters]


# ---------------------------------------------------------------------------
# OCR — heavily optimized for engineering legend text
# ---------------------------------------------------------------------------

def preprocess_for_ocr(img_bgr: np.ndarray, upscale_factor: int = OCR_UPSCALE_FACTOR) -> np.ndarray:
    """
    Aggressively preprocess an image ROI for maximum OCR accuracy.
    
    Pipeline:
    1. Upscale the ROI significantly (3x default) for sub-pixel clarity
    2. Convert to grayscale
    3. Apply bilateral filter (denoise while preserving edges)
    4. Sharpen the text
    5. Apply Otsu's thresholding for clean binarization
    6. Morphological cleanup to remove noise and strengthen text
    7. Add white border padding for Tesseract
    """
    h, w = img_bgr.shape[:2]
    
    # 1. Upscale for better OCR resolution
    if upscale_factor > 1:
        new_w = w * upscale_factor
        new_h = h * upscale_factor
        img_bgr = cv2.resize(img_bgr, (new_w, new_h), interpolation=cv2.INTER_CUBIC)
    
    # 2. Convert to grayscale
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    
    # 3. Bilateral filter — removes noise while keeping text edges sharp
    gray = cv2.bilateralFilter(gray, 9, 75, 75)
    
    # 4. Sharpen with unsharp mask
    blurred = cv2.GaussianBlur(gray, (0, 0), 3)
    sharpened = cv2.addWeighted(gray, 1.5, blurred, -0.5, 0)
    
    # 5. Otsu's binarization — automatically finds optimal threshold
    _, thresh = cv2.threshold(sharpened, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # 6. Ensure black text on white background
    if np.mean(thresh) < 127:
        thresh = cv2.bitwise_not(thresh)
    
    # 7. Morphological cleanup — remove small noise dots
    kernel_clean = np.ones((2, 2), np.uint8)
    thresh = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel_clean, iterations=1)
    
    # 8. Slight dilation to strengthen thin text strokes
    kernel_dilate = np.ones((2, 2), np.uint8)
    thresh = cv2.dilate(thresh, kernel_dilate, iterations=1)
    
    # 9. Add white border padding — Tesseract works better with border
    border_size = 20
    thresh = cv2.copyMakeBorder(
        thresh, border_size, border_size, border_size, border_size,
        cv2.BORDER_CONSTANT, value=255
    )
    
    return thresh


def ocr_roi_multi(img_bgr: np.ndarray) -> str:
    """
    Run OCR with multiple strategies and pick the best result.
    
    Strategies:
    - PSM 6 (uniform block of text) — best for multi-line descriptions
    - PSM 7 (single text line) — best for short single-line descriptions
    - PSM 4 (single column of text) — good for narrow columns
    
    The result with highest average character confidence is chosen.
    """
    preprocessed = preprocess_for_ocr(img_bgr)
    
    psm_modes = [6, 7, 4]
    best_text = ""
    best_conf = -1
    
    for psm in psm_modes:
        config = f"--oem 3 --psm {psm}"
        try:
            # Get text with confidence data
            data = pytesseract.image_to_data(
                preprocessed, config=config, output_type=pytesseract.Output.DICT
            )
            
            # Build text from words with their confidences
            words = []
            confidences = []
            for i, word in enumerate(data['text']):
                word = word.strip()
                conf = int(data['conf'][i])
                if word and conf > 0:  # Only consider words with positive confidence
                    words.append(word)
                    confidences.append(conf)
            
            text = " ".join(words)
            avg_conf = sum(confidences) / len(confidences) if confidences else 0
            
            if avg_conf > best_conf and text.strip():
                best_conf = avg_conf
                best_text = text
                
        except Exception:
            continue
    
    # Fallback: simple string extraction if confidence-based approach fails
    if not best_text.strip():
        try:
            config = "--oem 3 --psm 6"
            best_text = pytesseract.image_to_string(preprocessed, config=config)
        except Exception:
            pass
    
    # Clean up: collapse whitespace, strip
    best_text = " ".join(best_text.split())
    return best_text.strip()


def ocr_roi(img_bgr: np.ndarray) -> str:
    """Run Tesseract on a BGR ROI and return cleaned, corrected text."""
    raw_text = ocr_roi_multi(img_bgr)
    corrected = apply_ocr_corrections(raw_text)
    return corrected


# ---------------------------------------------------------------------------
# Symbol tight-crop
# ---------------------------------------------------------------------------

def tight_crop_symbol(cell_bgr: np.ndarray, padding: int = SYMBOL_PADDING) -> np.ndarray | None:
    """
    Find bounding box of non-white content and return tight crop with padding.
    """
    gray = cv2.cvtColor(cell_bgr, cv2.COLOR_BGR2GRAY)
    _, mask = cv2.threshold(gray, 235, 255, cv2.THRESH_BINARY_INV)

    # Erode away thin grid remnants at edges
    border = 3
    mask[:border, :] = 0
    mask[-border:, :] = 0
    mask[:, :border] = 0
    mask[:, -border:] = 0

    coords = cv2.findNonZero(mask)
    if coords is None:
        return None

    x, y, w, h = cv2.boundingRect(coords)
    H, W = cell_bgr.shape[:2]

    x1 = max(x - padding, 0)
    y1 = max(y - padding, 0)
    x2 = min(x + w + padding, W)
    y2 = min(y + h + padding, H)

    cropped = cell_bgr[y1:y2, x1:x2]
    return cropped if cropped.size > 0 else None


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

# Target icon size in Excel (in pixels). Images are scaled to fit within this.
EXCEL_ICON_MAX_WIDTH = 80
EXCEL_ICON_MAX_HEIGHT = 60


def export_to_excel(results: list[tuple[str, str]], output_path: str) -> None:
    """
    Export extracted icons and their names to a styled Excel workbook.

    Horizontal layout (1st row across multiple columns):
      Row 1-2:  Title (merged across all columns)
      Row 3  :  Description names — one per column
      Row 4  :  Icon images       — one per column

    Parameters
    ----------
    results : list of (description_text, icon_file_path) tuples
    output_path : path for the .xlsx file to create
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Legend Icons"

    # --- Styles ---
    header_font = Font(name="Calibri", size=14, bold=True)
    name_font = Font(name="Calibri", size=10, bold=True)
    name_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    img_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font_white = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill_light = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    alt_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Filter valid results first
    valid_results = [(desc, path) for desc, path in results if os.path.isfile(path)]
    num_icons = len(valid_results)

    if num_icons == 0:
        print("  [WARN] No valid icon files found for Excel export.")
        wb.save(output_path)
        return

    # --- Title row (merged across all icon columns) ---
    last_col_letter = get_column_letter(num_icons)
    ws.merge_cells(f"A1:{last_col_letter}2")
    title_cell = ws["A1"]
    title_cell.value = "LEGEND ICONS"
    title_cell.font = header_font_white
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = header_fill
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # Row heights for name and icon rows
    NAME_ROW = 4  # Row for description names
    ICON_ROW = 5  # Row for icon images
    ws.row_dimensions[NAME_ROW].height = 25
    ws.row_dimensions[ICON_ROW].height = 70  # enough for the icon image

    for col_idx, (desc_text, icon_path) in enumerate(valid_results, start=1):
        col_letter = get_column_letter(col_idx)

        # Alternating column colors for visual clarity
        bg_fill = alt_fill_light if (col_idx - 1) % 2 == 0 else alt_fill_white

        # Set column width (enough for icon + text)
        ws.column_dimensions[col_letter].width = 22

        # --- Name cell (row 4) ---
        name_cell = ws.cell(row=NAME_ROW, column=col_idx)
        name_cell.value = desc_text
        name_cell.font = name_font
        name_cell.alignment = name_alignment
        name_cell.fill = bg_fill
        name_cell.border = thin_border

        # --- Icon cell (row 5) ---
        icon_cell = ws.cell(row=ICON_ROW, column=col_idx)
        icon_cell.alignment = img_alignment
        icon_cell.fill = bg_fill
        icon_cell.border = thin_border

        # Insert the icon image
        try:
            img = XlImage(icon_path)

            # Scale image to fit within max dimensions while maintaining aspect ratio
            orig_w = img.width
            orig_h = img.height
            if orig_w > 0 and orig_h > 0:
                scale_w = EXCEL_ICON_MAX_WIDTH / orig_w
                scale_h = EXCEL_ICON_MAX_HEIGHT / orig_h
                scale = min(scale_w, scale_h, 1.0)  # don't upscale tiny icons beyond max
                if scale < 1.0:
                    img.width = int(orig_w * scale)
                    img.height = int(orig_h * scale)

            # Anchor image to the icon cell
            cell_ref = f"{col_letter}{ICON_ROW}"
            ws.add_image(img, cell_ref)
        except Exception as e:
            print(f"  [WARN] Could not embed icon '{icon_path}': {e}")
            icon_cell.value = f"[Image: {os.path.basename(icon_path)}]"

    # Save
    wb.save(output_path)
    print(f"[INFO] Excel file saved: {output_path}")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_legend(image_path: str, tesseract_cmd: str = None,
                   output_excel: str = None) -> list[tuple[str, str]]:
    """
    Process a legend table image: extract icons, OCR names, save PNGs, export Excel.

    Parameters
    ----------
    image_path : path to the input legend image
    tesseract_cmd : optional path to the Tesseract executable
    output_excel : optional path to save the Excel file; if None, auto-generated

    Returns
    -------
    list of (description_text, icon_file_path) tuples
    """
    if tesseract_cmd and os.path.isfile(tesseract_cmd):
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    if not os.path.isfile(image_path):
        raise FileNotFoundError(f"[ERROR] File not found: {image_path}")

    img_orig = cv2.imread(image_path)
    if img_orig is None:
        raise ValueError(f"[ERROR] Could not read image: {image_path}")

    print(f"[INFO] Original image size: {img_orig.shape[1]}x{img_orig.shape[0]}")

    # --- Upscale small images for much better OCR and line detection ---
    img, scale = upscale_image(img_orig, TARGET_WIDTH)
    print(f"[INFO] Working image size: {img.shape[1]}x{img.shape[0]}  (scale={scale:.2f}x)")

    # 1. Build line mask & detect grid coordinates
    line_mask = build_line_mask(img)
    h_coords, v_coords = detect_lines(line_mask)

    print(f"[INFO] Detected {len(h_coords)} horizontal lines at Y: {h_coords}")
    print(f"[INFO] Detected {len(v_coords)} vertical lines at X: {v_coords}")

    if len(h_coords) < 2:
        raise RuntimeError("[ERROR] Not enough horizontal lines detected.")
    if len(v_coords) < 2:
        raise RuntimeError("[ERROR] Not enough vertical lines detected.")

    # 2. Determine column divider
    left_x = v_coords[0]
    right_x = v_coords[-1]

    if len(v_coords) >= 3:
        table_w = right_x - left_x
        mid_target = left_x + table_w * 0.35  # symbol col is narrower (~35%)
        interior = v_coords[1:-1]
        col_divider = min(interior, key=lambda v: abs(v - mid_target))
    else:
        col_divider = (left_x + right_x) // 3  # assume symbol is ~1/3 width

    print(f"[INFO] Table: left={left_x}, divider={col_divider}, right={right_x}")

    # 3. Prepare output
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 4. Iterate rows — collect results
    used_names: dict[str, int] = {}
    results: list[tuple[str, str]] = []  # (description, icon_path)
    saved_count = 0
    margin = 4  # px margin inward past thick grid lines

    for idx in range(len(h_coords) - 1):
        y_top = h_coords[idx]
        y_bot = h_coords[idx + 1]
        row_height = y_bot - y_top

        if row_height < MIN_ROW_HEIGHT:
            print(f"  [SKIP] Row {idx}: too short ({row_height}px)")
            continue

        # --- Extract ROIs ---
        sym_roi = img[y_top + margin: y_bot - margin,
                      left_x + margin: col_divider - margin]
        desc_roi = img[y_top + margin: y_bot - margin,
                       col_divider + margin: right_x - margin]

        if sym_roi.size == 0 or desc_roi.size == 0:
            print(f"  [SKIP] Row {idx}: empty ROI")
            continue

        # --- OCR on description column ---
        desc_text = ""
        try:
            desc_text = ocr_roi(desc_roi)
        except Exception as e:
            print(f"  [WARN] OCR failed for row {idx}: {e}")

        # Also try OCR on the full row for header detection (LEGEND spans full width)
        full_row_roi = img[y_top + margin: y_bot - margin,
                           left_x + margin: right_x - margin]
        full_text = ""
        try:
            full_text = ocr_roi(full_row_roi)
        except Exception:
            pass

        # --- Skip header rows ---
        combined_lower = (desc_text + " " + full_text).lower()
        if any(kw in combined_lower for kw in HEADER_KEYWORDS):
            print(f"  [SKIP] Header row {idx}: \"{desc_text}\" / \"{full_text}\"")
            continue

        # --- Build filename ---
        sanitized = sanitize_filename(desc_text)
        if not sanitized:
            sanitized = f"row_{idx}"
            print(f"  [WARN] OCR empty for row {idx}; using fallback name.")
        
        final_name = unique_filename(sanitized, used_names)

        # --- Tight-crop the symbol ---
        cropped = tight_crop_symbol(sym_roi)
        if cropped is None:
            print(f"  [SKIP] Row {idx}: no symbol content detected.")
            continue

        # --- Save PNG ---
        out_path = os.path.join(OUTPUT_DIR, f"{final_name}.png")
        cv2.imwrite(out_path, cropped)
        saved_count += 1
        results.append((desc_text, os.path.abspath(out_path)))
        print(f"  [SAVE] Row {idx}: \"{desc_text}\"  ->  {out_path}")

    print(f"\n[DONE] Saved {saved_count} symbol(s) to '{OUTPUT_DIR}/'.")

    # 5. Export to Excel
    if results:
        if not output_excel:
            base_name = os.path.splitext(os.path.basename(image_path))[0]
            output_excel = os.path.join(_SCRIPT_DIR, f"{base_name}_legend_icons.xlsx")
        try:
            export_to_excel(results, output_excel)
            print(f"[DONE] Excel exported: {output_excel}")
        except Exception as e:
            print(f"[ERROR] Excel export failed: {e}")
    else:
        print("[WARN] No icons extracted — skipping Excel export.")

    return results


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python legend_processor.py <image_path>")
        print("Example: python legend_processor.py legend.jpeg")
        sys.exit(1)

    process_legend(sys.argv[1])
