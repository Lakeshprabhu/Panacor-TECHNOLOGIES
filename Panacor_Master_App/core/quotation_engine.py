import os
import sys
import json
import argparse
import google.generativeai as genai
from anthropic import Anthropic
from dotenv import load_dotenv
from pdf2image import convert_from_path
import base64
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

load_dotenv()

def convert_excel_to_csv(path):
    wb = load_workbook(path)
    ws = wb.active
    lines = []
    for row in ws.iter_rows():
        cells = []
        for cell in row:
            ref = f"{get_column_letter(cell.column)}{cell.row}"
            val = cell.value if cell.value is not None else ""
            cells.append(f"{ref}:{val}")
        lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)

def convert_image_to_base64(buffer):
    return base64.b64encode(buffer.getvalue()).decode("utf-8")

def convert_pdf_to_images(pdf_paths, dpi=150):
    """Convert PDFs to PIL images for Gemini."""
    all_images = []
    if isinstance(pdf_paths, str):
        pdf_paths = [pdf_paths]
    for pdf_path in pdf_paths:
        try:
            images = convert_from_path(pdf_path, dpi=dpi)
            all_images.extend(images)
        except Exception as e:
            print(f"Error converting PDF {pdf_path}: {e}")
    return all_images

def convert_pdf_to_base64(pdf_paths, dpi=100):
    """Convert PDFs to base64 images for Claude with prompt caching."""
    base64_array = []
    if isinstance(pdf_paths, str):
        pdf_paths = [pdf_paths]
    for pdf_path in pdf_paths:
        try:
            images = convert_from_path(pdf_path, dpi=dpi)
            for image in images:
                buffer = BytesIO()
                image.save(buffer, format="PNG")
                base64_array.append(convert_image_to_base64(buffer))
        except Exception as e:
            print(f"Error converting PDF {pdf_path}: {e}")
    return base64_array

# --- PROMPT TEMPLATES ---
PROMPT_DESCRIBE = """You are a data entry assistant. Here is an Excel file (formatted as CellReference:Value) and images from a PDF.
Describe the PDF data in a way that would allow you to fill in the empty cells in the Excel sheet.
Only mention the tables and information that would fit with the structure implied by the Excel data."""

PROMPT_EXTRACT = """Now fill in the EMPTY cells using information from the PDF. 
Rules:
1. Only fill cells that are currently empty or clearly placeholders.
2. Do NOT change existing valid data.
3. Return ONLY valid JSON array of changes, no other text.
4. Format: [{"cell": "B2", "value": "filled data", "type": "text"}, {"cell": "C3", "value": "42", "type": "number"}]
5. Types: "text", "number", "date"."""

def get_base_prompt(csv_data, user_prompt):
    """Legacy single-pass prompt for Gemini."""
    return f"""
    You are a data entry assistant.
    Here is data from an Excel file (formatted as CellReference:Value) and images from a PDF.
    Your task is to fill in the EMPTY cells in the Excel file using information found in the PDF images.
    
    Rules:
    1. Only fill cells that are currently empty or clearly placeholders.
    2. Do NOT change existing valid data.
    3. Return ONLY valid JSON array of changes.
    4. Format: [{{"cell": "B2", "value": "123.45", "type": "number"}}, ...]
    5. Types: "text", "number", "date".
    
    Current Excel Data:
    {csv_data}
    
    User Instructions: {user_prompt}
    """

# --- GEMINI IMPLEMENTATION ---
def run_gemini(api_key, images, full_prompt_text):
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel('gemini-1.5-flash')
    
    # Gemini accepts mixed list of text and PIL images
    content = [full_prompt_text, "\nPDF Images:"]
    content.extend(images)
    
    print("Sending to Gemini...")
    response = model.generate_content(content)
    return response.text

# --- CLAUDE IMPLEMENTATION WITH PROMPT CACHING ---
def run_claude_with_caching(api_key, base64_images, csv_data, user_prompt=""):
    """
    Two-phase Claude extraction with prompt caching.
    Phase 1: Describe the PDF data relevant to Excel structure
    Phase 2: Extract and fill cells with JSON output
    """
    client = Anthropic(api_key=api_key)
    
    # Build initial content with images (last image gets cache control)
    content = []
    for i, image_hash in enumerate(base64_images):
        image_block = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": image_hash
            }
        }
        # Add cache control to the last image for prompt caching
        if i == len(base64_images) - 1:
            image_block["cache_control"] = {"type": "ephemeral"}
        content.append(image_block)
    
    # Add Excel data
    content.append({"type": "text", "text": csv_data})
    
    # Add description prompt with user instructions
    describe_prompt = PROMPT_DESCRIBE
    if user_prompt:
        describe_prompt += f"\n\nAdditional instructions: {user_prompt}"
    content.append({"type": "text", "text": describe_prompt})
    
    messages = [{"role": "user", "content": content}]
    
    # Phase 1: Description
    print("Claude Phase 1: Analyzing PDF structure...")
    response1 = client.messages.create(
        max_tokens=4096,
        messages=messages,
        model="claude-haiku-4-5-20251001",
    )
    print(f"[tokens] in: {response1.usage.input_tokens}, out: {response1.usage.output_tokens}, "
          f"cache_create: {getattr(response1.usage, 'cache_creation_input_tokens', 0)}, "
          f"cache_read: {getattr(response1.usage, 'cache_read_input_tokens', 0)}")
    
    description = response1.content[0].text
    print(f"Description: {description[:200]}...")
    
    # Add assistant response and extraction prompt
    messages.append({"role": "assistant", "content": description})
    messages.append({"role": "user", "content": PROMPT_EXTRACT})
    
    # Phase 2: Extraction
    print("Claude Phase 2: Extracting data...")
    response2 = client.messages.create(
        max_tokens=4096,
        messages=messages,
        model="claude-haiku-4-5-20251001",
    )
    print(f"[tokens] in: {response2.usage.input_tokens}, out: {response2.usage.output_tokens}, "
          f"cache_create: {getattr(response2.usage, 'cache_creation_input_tokens', 0)}, "
          f"cache_read: {getattr(response2.usage, 'cache_read_input_tokens', 0)}")
    
    return response2.content[0].text

# --- LEGACY CLAUDE (single pass, no caching) ---
def run_claude(api_key, images, full_prompt_text):
    """Legacy single-pass Claude for backwards compatibility."""
    client = Anthropic(api_key=api_key)
    
    content_payload = []
    for img in images:
        buf = BytesIO()
        img.save(buf, format="PNG")
        b64_str = convert_image_to_base64(buf)
        content_payload.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": b64_str
            }
        })
    
    content_payload.append({
        "type": "text",
        "text": full_prompt_text
    })

    print("Sending to Claude (legacy mode)...")
    message = client.messages.create(
        max_tokens=4096,
        messages=[{"role": "user", "content": content_payload}],
        model="claude-haiku-4-5-20251001", 
    )
    return message.content[0].text

def process_quotation(excel_path, pdf_paths, user_prompt="", api_key=None, provider="Gemini"):
    """
    Main function to process quotation.
    provider: "Gemini" or "Claude"
    """
    # 1. Prepare Data
    print("Reading Excel...")
    csv_data = convert_excel_to_csv(excel_path)
    
    raw_response = ""
    
    if provider == "Gemini":
        # Gemini: Use PIL images and single-pass prompt
        print("Converting PDFs for Gemini...")
        images = convert_pdf_to_images(pdf_paths, dpi=150)
        prompt_text = get_base_prompt(csv_data, user_prompt)
        
        if not api_key: 
            api_key = os.getenv("GEMINI_API_KEY")
        if not api_key: 
            raise ValueError("Gemini API Key missing")
        raw_response = run_gemini(api_key, images, prompt_text)
        
    elif provider == "Claude":
        # Claude: Use two-phase extraction with prompt caching
        print("Converting PDFs for Claude (with caching)...")
        base64_images = convert_pdf_to_base64(pdf_paths, dpi=100)
        
        if not api_key: 
            api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key: 
            raise ValueError("Anthropic API Key missing")
        raw_response = run_claude_with_caching(api_key, base64_images, csv_data, user_prompt)
        
    else:
        raise ValueError(f"Unknown provider: {provider}")

    # 2. Parse JSON
    try:
        start = raw_response.find("[")
        end = raw_response.rfind("]")
        if start != -1 and end != -1:
            json_str = raw_response[start:end+1]
        else:
            json_str = raw_response
        changes = json.loads(json_str)
    except Exception as e:
        print("Raw Response:", raw_response)
        raise ValueError(f"Failed to parse AI response: {e}")

    # 3. Apply Changes
    print(f"Applying {len(changes)} changes...")
    wb = load_workbook(excel_path)
    ws = wb.active

    for change in changes:
        val = change["value"]
        cell_ref = change["cell"]
        if change.get("type") == "number":
            try:
                clean_val = str(val).replace(",", "").replace("$", "").strip()
                f_val = float(clean_val)
                val = int(f_val) if f_val.is_integer() else f_val
            except ValueError:
                pass 
        ws[cell_ref] = val

    out_path = excel_path.rsplit(".", 1)[0] + "_filled.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("excel")
    parser.add_argument("pdfs", nargs="+")
    parser.add_argument("-p", "--prompt", default="")
    parser.add_argument("--key", help="API Key")
    parser.add_argument("--provider", default="Gemini", choices=["Gemini", "Claude"])
    args = parser.parse_args()

    try:
        out = process_quotation(args.excel, args.pdfs, args.prompt, args.key, args.provider)
        print(f"Success! Output: {out}")
    except Exception as e:
        print(f"Error: {e}")
