#!/usr/bin/env python3
"""
O&M Manual Preparation — Core Processor
=========================================
Reads items from various input formats, searches for technical datasheets
online, downloads them, and merges everything into a single PDF.

Supported input formats:
  1. PDF  — extracts text lines
  2. Excel — reads non-empty cells
  3. Image — OCR with Tesseract
  4. Text file — reads lines
  5. Manual text — split by newlines
"""

import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime

import cv2
import fitz  # PyMuPDF
import numpy as np
import pytesseract
import requests
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(_SCRIPT_DIR, "downloaded_datasheets")
CACHE_FILE = os.path.join(_SCRIPT_DIR, ".search_cache.json")

TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.isfile(TESSERACT_CMD):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

# Search / download settings
SEARCH_DELAY_SEC = 2.5          # seconds between Google searches
DOWNLOAD_TIMEOUT = 30           # seconds per download
MAX_SEARCH_RESULTS = 10         # results to inspect per item
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)

# Known datasheet / technical-doc hosting domains (boost score)
KNOWN_DATASHEET_DOMAINS = {
    "mouser.com", "digikey.com", "rs-online.com", "farnell.com",
    "alldatasheet.com", "datasheetcatalog.com", "octopart.com",
    "arrow.com", "newark.com", "element14.com", "ti.com",
    "st.com", "nxp.com", "microchip.com", "analog.com",
    "infineon.com", "onsemi.com", "vishay.com", "te.com",
    "schneider-electric.com", "siemens.com", "abb.com",
    "legrand.com", "honeywell.com", "carrier.com",
    "eurovent-certification.com",
}


# ---------------------------------------------------------------------------
# Search-result cache  (avoids re-searching the same item)
# ---------------------------------------------------------------------------

def _load_cache() -> dict:
    if os.path.isfile(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_cache(cache: dict) -> None:
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


def _cache_key(item_name: str) -> str:
    return hashlib.md5(item_name.strip().lower().encode()).hexdigest()


# ---------------------------------------------------------------------------
# 1.  Input parsers  — each returns list[str]
# ---------------------------------------------------------------------------

def parse_items_from_pdf(pdf_path: str) -> list[str]:
    """Extract non-empty text lines from every page of a PDF."""
    doc = fitz.open(pdf_path)
    items: list[str] = []
    for page in doc:
        text = page.get_text("text")
        for line in text.splitlines():
            cleaned = line.strip()
            if cleaned and len(cleaned) > 2:
                items.append(cleaned)
    doc.close()
    return _deduplicate(items)


def parse_items_from_excel(excel_path: str) -> list[str]:
    """Read every non-empty cell from all sheets of an Excel workbook."""
    wb = load_workbook(excel_path, data_only=True)
    items: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).strip()
                    if val and len(val) > 2:
                        items.append(val)
    wb.close()
    return _deduplicate(items)


def parse_items_from_image(image_path: str) -> list[str]:
    """OCR an image and return each detected text line."""
    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Could not read image: {image_path}")

    # Pre-process: grayscale → threshold for cleaner OCR
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    text = pytesseract.image_to_string(thresh, config="--oem 3 --psm 6")
    items: list[str] = []
    for line in text.splitlines():
        cleaned = line.strip()
        if cleaned and len(cleaned) > 2:
            items.append(cleaned)
    return _deduplicate(items)


def parse_items_from_text(text_path: str) -> list[str]:
    """Read a plain-text file line by line."""
    with open(text_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
    items: list[str] = []
    for line in lines:
        cleaned = line.strip()
        if cleaned and len(cleaned) > 2:
            items.append(cleaned)
    return _deduplicate(items)


def parse_items_from_string(text: str) -> list[str]:
    """Parse manually-typed text (newline-separated)."""
    items: list[str] = []
    for line in text.splitlines():
        cleaned = line.strip()
        if cleaned and len(cleaned) > 2:
            items.append(cleaned)
    return _deduplicate(items)


def _deduplicate(items: list[str]) -> list[str]:
    """Remove exact duplicates while preserving order."""
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        key = item.strip().lower()
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


# ---------------------------------------------------------------------------
# 2.  Web search  — find datasheet PDF URLs
# ---------------------------------------------------------------------------

def _score_url(url: str) -> int:
    """Score a URL for datasheet relevance. Higher = better."""
    score = 0
    url_lower = url.lower()

    if url_lower.endswith(".pdf"):
        score += 10
    if "datasheet" in url_lower:
        score += 5
    if "technical" in url_lower or "spec" in url_lower:
        score += 3
    if "manual" in url_lower or "brochure" in url_lower:
        score += 2

    for domain in KNOWN_DATASHEET_DOMAINS:
        if domain in url_lower:
            score += 3
            break

    return score


def search_datasheet(item_name: str, num_results: int = MAX_SEARCH_RESULTS) -> str | None:
    """
    Search Google for a technical datasheet PDF for the given item.

    Returns the best-scoring URL, or None if nothing found.
    Uses googlesearch-python library.
    """
    try:
        from googlesearch import search as gsearch
    except ImportError:
        print("[ERROR] googlesearch-python not installed. Run: pip install googlesearch-python")
        return None

    query = f"{item_name} technical datasheet filetype:pdf"
    print(f"  [SEARCH] Querying: {query}")

    try:
        results = list(gsearch(query, num_results=num_results))
    except Exception as e:
        print(f"  [WARN] Google search failed: {e}")
        # Try a simpler query as fallback
        try:
            query2 = f"{item_name} datasheet PDF"
            print(f"  [SEARCH] Retry with: {query2}")
            results = list(gsearch(query2, num_results=num_results))
        except Exception as e2:
            print(f"  [ERROR] Search failed completely: {e2}")
            return None

    if not results:
        print(f"  [WARN] No results found for: {item_name}")
        return None

    # Score and sort results
    scored = [(url, _score_url(url)) for url in results]
    scored.sort(key=lambda x: x[1], reverse=True)

    best_url = scored[0][0]
    print(f"  [FOUND] Best URL (score={scored[0][1]}): {best_url}")
    return best_url


# ---------------------------------------------------------------------------
# 3.  PDF download & validation
# ---------------------------------------------------------------------------

def _is_valid_pdf(filepath: str) -> bool:
    """Check if a file starts with the PDF magic bytes."""
    try:
        with open(filepath, "rb") as f:
            header = f.read(5)
        return header == b"%PDF-"
    except Exception:
        return False


def download_pdf(url: str, save_path: str) -> bool:
    """
    Download a file from a URL and save it.
    Returns True if the downloaded file is a valid PDF.
    """
    try:
        headers = {"User-Agent": USER_AGENT}
        resp = requests.get(url, headers=headers, timeout=DOWNLOAD_TIMEOUT,
                            stream=True, allow_redirects=True)
        resp.raise_for_status()

        # Check content-type hint (not always reliable)
        content_type = resp.headers.get("Content-Type", "").lower()

        with open(save_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)

        if _is_valid_pdf(save_path):
            size_kb = os.path.getsize(save_path) / 1024
            print(f"  [DOWNLOAD] OK — {size_kb:.0f} KB saved: {os.path.basename(save_path)}")
            return True
        else:
            # Downloaded file is not a PDF (could be HTML login page, etc.)
            print(f"  [WARN] Downloaded file is not a valid PDF, removing: {save_path}")
            os.remove(save_path)
            return False

    except Exception as e:
        print(f"  [ERROR] Download failed for {url}: {e}")
        if os.path.isfile(save_path):
            os.remove(save_path)
        return False


def _safe_filename(item_name: str) -> str:
    """Convert an item name into a safe filename."""
    name = re.sub(r'[\\/:*?"<>|]+', '_', item_name.strip())
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    if len(name) > 80:
        name = name[:80]
    return name


# ---------------------------------------------------------------------------
# 4.  PDF merger with table of contents
# ---------------------------------------------------------------------------

def merge_pdfs(pdf_paths: list[tuple[str, str]], output_path: str) -> None:
    """
    Merge multiple PDFs into one, prepending a Table of Contents page.

    Parameters
    ----------
    pdf_paths : list of (item_name, pdf_file_path) tuples
    output_path : path for the combined output PDF
    """
    if not pdf_paths:
        print("[WARN] No PDFs to merge.")
        return

    merged = fitz.open()  # new empty PDF

    # ------ Build TOC data (item_name → page number) ---------------------
    toc_entries: list[tuple[str, int]] = []
    current_page = 1  # page 1 is the TOC itself

    # First, figure out how many pages each source has
    page_counts: list[int] = []
    for item_name, pdf_path in pdf_paths:
        try:
            src = fitz.open(pdf_path)
            page_counts.append(len(src))
            src.close()
        except Exception:
            page_counts.append(0)

    # TOC page is page 0 in 0-indexed; datasheets start at page 1
    datasheet_start_page = 1
    for idx, (item_name, pdf_path) in enumerate(pdf_paths):
        toc_entries.append((item_name, datasheet_start_page + 1))  # 1-indexed for display
        datasheet_start_page += page_counts[idx]

    # ------ Create TOC page (A4) ----------------------------------------
    toc_page_rect = fitz.paper_rect("a4")
    toc_doc = fitz.open()
    toc_page = toc_doc.new_page(width=toc_page_rect.width, height=toc_page_rect.height)

    # Title
    title_point = fitz.Point(50, 60)
    toc_page.insert_text(title_point, "TABLE OF CONTENTS",
                         fontsize=20, fontname="helv", color=(0.18, 0.33, 0.59))

    # Underline
    toc_page.draw_line(fitz.Point(50, 70), fitz.Point(545, 70),
                       color=(0.18, 0.33, 0.59), width=1.5)

    # Entries
    y_pos = 100
    for i, (item_name, page_num) in enumerate(toc_entries, start=1):
        if y_pos > 780:
            # Add a new TOC page if we run out of space
            toc_page = toc_doc.new_page(width=toc_page_rect.width,
                                         height=toc_page_rect.height)
            y_pos = 60

        entry_text = f"{i}.  {item_name}"
        page_text = f"Page {page_num}"

        toc_page.insert_text(fitz.Point(60, y_pos), entry_text,
                             fontsize=10, fontname="helv", color=(0, 0, 0))
        toc_page.insert_text(fitz.Point(480, y_pos), page_text,
                             fontsize=10, fontname="helv", color=(0.4, 0.4, 0.4))

        # Dotted leader line
        leader_x_start = 60 + len(entry_text) * 4.5
        leader_x_start = min(leader_x_start, 470)
        for lx in range(int(leader_x_start), 475, 5):
            toc_page.draw_circle(fitz.Point(lx, y_pos - 2), 0.5,
                                 color=(0.7, 0.7, 0.7), fill=(0.7, 0.7, 0.7))

        y_pos += 20

    # Insert TOC pages into merged doc
    merged.insert_pdf(toc_doc)
    toc_doc.close()

    # ------ Append each datasheet ----------------------------------------
    for item_name, pdf_path in pdf_paths:
        try:
            src = fitz.open(pdf_path)
            merged.insert_pdf(src)
            src.close()
            print(f"  [MERGE] Added: {item_name} ({os.path.basename(pdf_path)})")
        except Exception as e:
            print(f"  [ERROR] Could not merge '{pdf_path}': {e}")

    # ------ Save ---------------------------------------------------------
    merged.save(output_path)
    merged.close()
    print(f"[INFO] Merged PDF saved: {output_path}")


# ---------------------------------------------------------------------------
# 5.  Main pipeline
# ---------------------------------------------------------------------------

def fetch_all_datasheets(
    items: list[str],
    progress_callback=None,
) -> list[tuple[str, str]]:
    """
    Search and download datasheets for all items.

    Parameters
    ----------
    items : list of item names
    progress_callback : optional callable(current_index, total, item_name, status)

    Returns
    -------
    list of (item_name, downloaded_pdf_path) for successful downloads
    """
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    cache = _load_cache()
    downloaded: list[tuple[str, str]] = []
    total = len(items)

    for idx, item_name in enumerate(items):
        if progress_callback:
            progress_callback(idx, total, item_name, "searching")

        key = _cache_key(item_name)
        safe_name = _safe_filename(item_name)
        save_path = os.path.join(DOWNLOAD_DIR, f"{safe_name}.pdf")

        # Check cache — skip if already downloaded and file still exists
        if key in cache and os.path.isfile(cache[key].get("path", "")):
            cached_path = cache[key]["path"]
            if _is_valid_pdf(cached_path):
                print(f"[{idx+1}/{total}] CACHED: {item_name}")
                downloaded.append((item_name, cached_path))
                if progress_callback:
                    progress_callback(idx, total, item_name, "cached")
                continue

        print(f"\n[{idx+1}/{total}] Processing: {item_name}")

        # Search
        url = search_datasheet(item_name)

        if url:
            if progress_callback:
                progress_callback(idx, total, item_name, "downloading")

            success = download_pdf(url, save_path)

            if success:
                downloaded.append((item_name, save_path))
                cache[key] = {
                    "item": item_name,
                    "url": url,
                    "path": save_path,
                    "timestamp": datetime.now().isoformat(),
                }
                _save_cache(cache)
            else:
                print(f"  [SKIP] Could not download a valid PDF for: {item_name}")
        else:
            print(f"  [SKIP] No datasheet found for: {item_name}")

        if progress_callback:
            status = "done" if url else "not_found"
            progress_callback(idx, total, item_name, status)

        # Rate limiting — wait before next search
        if idx < total - 1:
            time.sleep(SEARCH_DELAY_SEC)

    return downloaded


def process_om_manual(
    items: list[str],
    output_pdf: str | None = None,
    progress_callback=None,
) -> str | None:
    """
    Full pipeline: search → download → merge.

    Parameters
    ----------
    items : list of item names
    output_pdf : optional output path; auto-generated if None
    progress_callback : optional progress callable

    Returns
    -------
    Path to the merged PDF, or None if nothing was produced.
    """
    if not items:
        print("[WARN] No items to process.")
        return None

    print(f"[INFO] Processing {len(items)} item(s)...")

    # Step 1: Search & download
    downloaded = fetch_all_datasheets(items, progress_callback)

    if not downloaded:
        print("[WARN] No datasheets were downloaded. Cannot produce a merged PDF.")
        return None

    print(f"\n[INFO] Downloaded {len(downloaded)}/{len(items)} datasheet(s).")

    # Step 2: Merge
    if not output_pdf:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_pdf = os.path.join(_SCRIPT_DIR, f"OM_Manual_{timestamp}.pdf")

    merge_pdfs(downloaded, output_pdf)
    print(f"\n[DONE] O&M Manual created: {output_pdf}")
    return output_pdf


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python om_processor.py <input_file> [--type pdf|excel|image|text]")
        print("       python om_processor.py --manual \"Item 1\" \"Item 2\" ...")
        sys.exit(1)

    if sys.argv[1] == "--manual":
        item_list = sys.argv[2:]
    else:
        input_file = sys.argv[1]
        file_type = "text"  # default

        # Auto-detect from extension
        ext = os.path.splitext(input_file)[1].lower()
        if ext == ".pdf":
            file_type = "pdf"
        elif ext in (".xlsx", ".xls"):
            file_type = "excel"
        elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff"):
            file_type = "image"

        # Override with --type flag
        if "--type" in sys.argv:
            ti = sys.argv.index("--type")
            if ti + 1 < len(sys.argv):
                file_type = sys.argv[ti + 1]

        parsers = {
            "pdf": parse_items_from_pdf,
            "excel": parse_items_from_excel,
            "image": parse_items_from_image,
            "text": parse_items_from_text,
        }

        parser = parsers.get(file_type)
        if not parser:
            print(f"[ERROR] Unknown type: {file_type}")
            sys.exit(1)

        print(f"[INFO] Parsing {file_type} file: {input_file}")
        item_list = parser(input_file)

    print(f"[INFO] Found {len(item_list)} item(s):")
    for i, item in enumerate(item_list, 1):
        print(f"  {i}. {item}")

    process_om_manual(item_list)
