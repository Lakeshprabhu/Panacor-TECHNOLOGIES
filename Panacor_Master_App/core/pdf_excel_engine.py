#!/usr/bin/env python3
"""
PDF to Excel Extractor — PANACOR TECHNOLOGIES
==============================================
Extracts text from a PDF file line-by-line and outputs a formatted
Excel workbook with:
  - Row 1: PDF filename as title (merged across all columns)
  - Column 1 (S.No): Auto-incremented serial number
  - Column 2 (Description): Extracted text lines
  - Additional user-defined columns (variable count & names)
  - Bold highlighting for headings / section titles
  - Full bordered table styling

Usage:
    python pdf_extractor.py <pdf_path>
"""

import os
import re
import sys

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


# ---------------------------------------------------------------------------
# PDF text extraction with bold / heading detection
# ---------------------------------------------------------------------------

def extract_lines_from_pdf(pdf_path: str) -> list[dict]:
    """
    Extract text from every page of a PDF.

    Returns a list of dicts, one per logical line:
        {
            "text":      str,   # the line content
            "is_bold":   bool,  # True if the dominant font is bold / the line is a heading
            "page":      int,   # 1-based page number
        }

    Heading detection heuristics:
      1. Font flags contain bold (bit 4 in fitz span flags)
      2. Font size is noticeably larger than the document median
      3. The line is ALL-CAPS and short (≤ 80 chars)
    """
    doc = fitz.open(pdf_path)
    raw_spans: list[dict] = []          # every span across all pages
    all_font_sizes: list[float] = []    # to compute the median later

    # ------ Pass 1: collect every span ----------------------------------
    for page_num in range(len(doc)):
        page = doc[page_num]
        blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]

        for block in blocks:
            if block["type"] != 0:          # skip images / drawings
                continue
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"]
                    if not text.strip():
                        continue
                    size = span["size"]
                    flags = span["flags"]   # bit-field; bit 4 = bold
                    is_bold_font = bool(flags & (1 << 4))
                    raw_spans.append({
                        "text": text,
                        "size": size,
                        "is_bold_font": is_bold_font,
                        "page": page_num + 1,
                        "block_no": block["number"],
                        "line_y": line["bbox"][1],      # vertical position
                    })
                    all_font_sizes.append(size)

    doc.close()

    if not raw_spans:
        return []

    # ------ Compute median font size ------------------------------------
    all_font_sizes.sort()
    median_size = all_font_sizes[len(all_font_sizes) // 2]

    # ------ Pass 2: merge spans into logical lines ----------------------
    lines: list[dict] = []
    current_line_parts: list[dict] = [raw_spans[0]]

    for span in raw_spans[1:]:
        prev = current_line_parts[-1]
        # Same page, same block, similar Y → same line
        same_line = (
            span["page"] == prev["page"]
            and span["block_no"] == prev["block_no"]
            and abs(span["line_y"] - prev["line_y"]) < 2
        )
        if same_line:
            current_line_parts.append(span)
        else:
            lines.append(_merge_spans(current_line_parts, median_size))
            current_line_parts = [span]

    # flush last line
    if current_line_parts:
        lines.append(_merge_spans(current_line_parts, median_size))

    return lines


def _merge_spans(spans: list[dict], median_size: float) -> dict:
    """Merge a list of spans that belong to the same visual line."""
    text = "".join(s["text"] for s in spans).strip()
    text = re.sub(r"[ \t]+", " ", text)  # collapse internal whitespace

    # Bold if majority of chars come from bold spans
    bold_chars = sum(len(s["text"]) for s in spans if s["is_bold_font"])
    total_chars = max(sum(len(s["text"]) for s in spans), 1)
    majority_bold = (bold_chars / total_chars) >= 0.5

    # Large-font heading: size ≥ 1.2× the median
    max_size = max(s["size"] for s in spans)
    is_large = max_size >= median_size * 1.2

    # ALL-CAPS short line heuristic
    is_allcaps_short = text.isupper() and len(text) <= 80

    is_bold = majority_bold or is_large or is_allcaps_short

    return {
        "text": text,
        "is_bold": is_bold,
        "page": spans[0]["page"],
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(
    lines: list[dict],
    output_path: str,
    pdf_filename: str,
    extra_columns: list[str] | None = None,
) -> None:
    """
    Write the extracted lines into a styled Excel workbook.

    Layout
    ------
    Row 1      : PDF filename (title), merged across all columns
    Row 2      : Column headers — S.No | Description | <extra columns …>
    Row 3+     : Data rows, one per extracted line

    Headings / bold lines get bold font styling in the Description cell.
    The entire sheet is formatted as a bordered table.

    Parameters
    ----------
    lines         : output of extract_lines_from_pdf()
    output_path   : .xlsx save path
    pdf_filename  : original PDF filename (shown in title row)
    extra_columns : optional list of additional column header names
    """
    if extra_columns is None:
        extra_columns = []

    all_headers = ["S.No", "Description"] + extra_columns
    num_cols = len(all_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # ── Styles ──────────────────────────────────────────────────────────
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    normal_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    sno_alignment = Alignment(horizontal="center", vertical="center")
    desc_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    extra_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    alt_fill_light = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    alt_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ── Row 1: Title (PDF filename) ─────────────────────────────────────
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = pdf_filename
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    title_cell.fill = title_fill
    ws.row_dimensions[1].height = 30

    # Apply title fill + border to every cell across the merged row
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = title_fill
        cell.border = thin_border

    # ── Row 2: Column headers ───────────────────────────────────────────
    HEADER_ROW = 2
    for col_idx, header_name in enumerate(all_headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.value = header_name
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border
    ws.row_dimensions[HEADER_ROW].height = 22

    # ── Column widths ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8      # S.No
    ws.column_dimensions["B"].width = 80     # Description
    for c in range(3, num_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ── Data rows ───────────────────────────────────────────────────────
    DATA_START = 3
    serial = 1

    for idx, line_info in enumerate(lines):
        row_num = DATA_START + idx
        text = line_info["text"]
        is_bold = line_info["is_bold"]
        bg_fill = alt_fill_light if idx % 2 == 0 else alt_fill_white

        # S.No
        sno_cell = ws.cell(row=row_num, column=1)
        sno_cell.value = serial
        sno_cell.font = bold_font if is_bold else normal_font
        sno_cell.alignment = sno_alignment
        sno_cell.fill = bg_fill
        sno_cell.border = thin_border

        # Description
        desc_cell = ws.cell(row=row_num, column=2)
        desc_cell.value = text
        desc_cell.font = bold_font if is_bold else normal_font
        desc_cell.alignment = desc_alignment
        desc_cell.fill = bg_fill
        desc_cell.border = thin_border

        # Extra columns (empty but styled)
        for c in range(3, num_cols + 1):
            extra_cell = ws.cell(row=row_num, column=c)
            extra_cell.font = normal_font
            extra_cell.alignment = extra_alignment
            extra_cell.fill = bg_fill
            extra_cell.border = thin_border

        serial += 1

    # ── Save ────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"[INFO] Excel file saved: {output_path}")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_pdf(
    pdf_path: str,
    output_excel: str | None = None,
    extra_columns: list[str] | None = None,
) -> list[dict]:
    """
    Full pipeline: extract PDF text → export to styled Excel.

    Parameters
    ----------
    pdf_path      : path to the input PDF
    output_excel  : optional .xlsx output path; auto-generated if None
    extra_columns : optional list of additional column names

    Returns
    -------
    list of extracted line dicts
    """
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"[ERROR] File not found: {pdf_path}")

    print(f"[INFO] Opening PDF: {pdf_path}")
    lines = extract_lines_from_pdf(pdf_path)
    print(f"[INFO] Extracted {len(lines)} lines from PDF.")

    if not lines:
        print("[WARN] No text found in the PDF.")
        return []

    # Count headings
    bold_count = sum(1 for l in lines if l["is_bold"])
    print(f"[INFO] Detected {bold_count} heading/bold line(s).")

    # Auto-generate output path
    if not output_excel:
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_excel = os.path.join(_SCRIPT_DIR, f"{base_name}_extracted.xlsx")

    pdf_filename = os.path.basename(pdf_path)
    export_to_excel(lines, output_excel, pdf_filename, extra_columns)
    print(f"[DONE] Excel exported: {output_excel}")

    return lines


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python pdf_extractor.py <pdf_path> [extra_col1] [extra_col2] ...")
        print("Example: python pdf_extractor.py document.pdf 'Unit' 'Qty' 'Rate'")
        sys.exit(1)

    pdf_file = sys.argv[1]
    extras = sys.argv[2:] if len(sys.argv) > 2 else []
    process_pdf(pdf_file, extra_columns=extras)
